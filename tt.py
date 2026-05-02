import asyncio
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import async_playwright
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes
import os, json, re
from telegram.request import HTTPXRequest

try:
    from config import USERNAME, TELEGRAM_TOKEN, CHAT_ID, PROXY, INTERVAL_HOURS, EXCEL_FILE, SEEN_FILE, DELETED_FILE, PROFILE_DIR
except ImportError:
    raise SystemExit(
        "❌ Файл config.py не найден!\n"
        "Скопируй config.example.py → config.py и заполни свои данные."
    )

next_check_time = None
is_checking = False

def load_seen() -> set:
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE) as f:
            return set(json.load(f))
    return set()

def save_seen(ids: set):
    with open(SEEN_FILE, "w") as f:
        json.dump(list(ids), f)

def load_deleted() -> dict:
    if os.path.exists(DELETED_FILE):
        with open(DELETED_FILE) as f:
            return json.load(f)
    return {}

def save_deleted(deleted: dict):
    with open(DELETED_FILE, "w") as f:
        json.dump(deleted, f, ensure_ascii=False)

async def _run_browser(username: str) -> list[dict]:
    results = []
    intercepted = []

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            headless=False,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        )

        page = await context.new_page()

        async def handle_response(response):
            try:
                if "repost" in response.url.lower():
                    data = await response.json()
                    intercepted.append(data)
            except:
                pass

        page.on("response", handle_response)

        await page.goto(f"https://www.tiktok.com/@{username}", wait_until="networkidle", timeout=60000)
        await asyncio.sleep(3)

        try:
            await page.wait_for_selector('[data-e2e="repost-tab"]', timeout=10000)
            tabs = await page.query_selector_all('[data-e2e="repost-tab"]')
            for tab in tabs:
                selected = await tab.get_attribute("aria-selected")
                if selected == "false":
                    await tab.click()
                    break
            await page.wait_for_selector('[data-e2e="repost-tab"][aria-selected="true"]', timeout=10000)
            print("Вкладка Репосты открыта")
            await asyncio.sleep(4)
        except Exception as e:
            print(f"Ошибка при переходе на вкладку репостов: {e}")

        for _ in range(8):
            await page.keyboard.press("End")
            await asyncio.sleep(2)

        for data in intercepted:
            items = []
            if isinstance(data, dict):
                items = (
                    data.get("itemList") or
                    data.get("items") or
                    data.get("data", {}).get("itemList") or
                    []
                )
            for item in items:
                try:
                    video_id = str(item.get("id", ""))
                    author = item.get("author", {}).get("uniqueId", "")
                    create_time = item.get("createTime", 0)
                    stats = item.get("stats", {})
                    results.append({
                        "id": video_id,
                        "desc": item.get("desc", ""),
                        "author": author,
                        "date": datetime.fromtimestamp(create_time).strftime("%Y-%m-%d %H:%M:%S") if create_time else "",
                        "likes": stats.get("diggCount", 0),
                        "shares": stats.get("shareCount", 0),
                        "url": f"https://tiktok.com/@{author}/video/{video_id}"
                    })
                except:
                    continue

        if not results:
            html = await page.content()
            with open("debug.html", "w", encoding="utf-8") as f:
                f.write(html)
            match = re.search(r'"itemList":\s*(\[.*?\])', html, re.DOTALL)
            if match:
                try:
                    items = json.loads(match.group(1))
                    for item in items:
                        video_id = str(item.get("id", ""))
                        author = item.get("author", {}).get("uniqueId", "")
                        results.append({
                            "id": video_id,
                            "desc": item.get("desc", ""),
                            "author": author,
                            "date": "",
                            "likes": 0,
                            "shares": 0,
                            "url": f"https://tiktok.com/@{author}/video/{video_id}"
                        })
                except:
                    pass

        await context.close()
    return results

async def get_reposts(username: str) -> list[dict]:
    try:
        return await _run_browser(username)
    except Exception as e:
        print(f"Ошибка браузера: {e}")
        return []

def init_excel(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Репосты"
    ws2 = wb.create_sheet("Удалённые")
    for sheet, headers in [
        (ws, ["ID видео", "Описание", "Автор", "Дата", "Лайки", "Шеры", "Ссылка", "Обнаружено"]),
        (ws2, ["ID видео", "Описание", "Автор", "Дата", "Лайки", "Шеры", "Ссылка", "Обнаружено", "Удалено"])
    ]:
        for col, h in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="4F81BD")
            cell.alignment = Alignment(horizontal="center")
        for col_letter, width in zip("ABCDEFGHI", [15, 40, 20, 20, 10, 10, 50, 20, 20]):
            sheet.column_dimensions[col_letter].width = width
    wb.save(path)

def save_new(reposts: list[dict], seen: set, path: str) -> tuple[set, list, list]:
    if not os.path.exists(path):
        init_excel(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["Репосты"]
    ws2 = wb["Удалённые"] if "Удалённые" in wb.sheetnames else wb.create_sheet("Удалённые")

    deleted_data = load_deleted()
    current_ids = {r["id"] for r in reposts if r["id"]}

    rows_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows_data[str(row[0])] = row

    newly_deleted = []
    for vid_id, row in rows_data.items():
        if vid_id in seen and vid_id not in current_ids and vid_id not in deleted_data:
            deleted_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            deleted_data[vid_id] = deleted_time
            ws2.append(list(row) + [deleted_time])
            newly_deleted.append({
                "id": vid_id,
                "author": row[2] if len(row) > 2 else "",
                "desc": row[1] if len(row) > 1 else "",
                "url": row[6] if len(row) > 6 else "",
                "deleted_at": deleted_time
            })
            for col_row in ws.iter_rows(min_row=2):
                if str(col_row[0].value) == vid_id:
                    for cell in col_row:
                        cell.fill = PatternFill("solid", start_color="FFCCCC")
                    break

    save_deleted(deleted_data)

    new_items = []
    for r in reposts:
        if r["id"] and r["id"] not in seen and r["id"] not in deleted_data:
            ws.append([
                r["id"], r["desc"], r["author"], r["date"],
                r["likes"], r["shares"], r["url"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
            seen.add(r["id"])
            new_items.append(r)

    wb.save(path)
    return seen, new_items, newly_deleted

async def do_check(app=None, reply_func=None):
    global seen, next_check_time, is_checking

    if is_checking:
        msg = "⏳ Проверка уже идёт, подожди..."
        if reply_func:
            await reply_func(msg)
        return

    is_checking = True
    try:
        reposts = await get_reposts(USERNAME)
        seen, new_items, deleted_items = save_new(reposts, seen, EXCEL_FILE)
        save_seen(seen)
        next_check_time = datetime.now().timestamp() + INTERVAL_HOURS * 3600

        text = ""
        if new_items:
            text += f"✅ Новых репостов: {len(new_items)}\n"
            for r in new_items[:10]:
                text += f"• [{r['author']}]({r['url']}) — {r['desc'][:50]}\n"
            if len(new_items) > 10:
                text += f"_...и ещё {len(new_items) - 10}_\n"

        if deleted_items:
            text += f"\n🗑 Удалено репостов: {len(deleted_items)}\n"
            for r in deleted_items[:10]:
                text += f"• [{r['author']}]({r['url']}) — {r['desc'][:50]}\n"
            if len(deleted_items) > 10:
                text += f"_...и ещё {len(deleted_items) - 10}_\n"

        if not text:
            text = "✅ Изменений нет."

        if reply_func:
            await reply_func(text, parse_mode="Markdown")
        elif app:
            await app.bot.send_message(CHAT_ID, text, parse_mode="Markdown")
    finally:
        is_checking = False

async def cmd_refresh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🔄 Запускаю проверку репостов...")
    asyncio.create_task(do_check(
        app=context.application,
        reply_func=update.message.reply_text
    ))

async def cmd_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [
            InlineKeyboardButton("30 сек (тест)", callback_data="time_0.008"),
            InlineKeyboardButton("30 мин", callback_data="time_0.5"),
        ],
        [
            InlineKeyboardButton("1 час", callback_data="time_1"),
            InlineKeyboardButton("2 часа", callback_data="time_2"),
        ],
        [
            InlineKeyboardButton("5 часов", callback_data="time_5"),
            InlineKeyboardButton("8 часов", callback_data="time_8"),
            InlineKeyboardButton("12 часов", callback_data="time_12"),
        ]
    ]
    await update.message.reply_text(
        "⏱ Выбери интервал проверки:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def callback_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global INTERVAL_HOURS, next_check_time
    query = update.callback_query
    await query.answer()
    hours = float(query.data.split("_")[1])
    INTERVAL_HOURS = hours
    next_check_time = datetime.now().timestamp() + hours * 3600
    label = {
        0.008: "30 секунд (тест)",
        0.5: "30 минут",
        1: "1 час",
        2: "2 часа",
        5: "5 часов",
        8: "8 часов",
        12: "12 часов"
    }
    await query.edit_message_text(f"✅ Интервал обновлён: {label.get(hours, f'{hours}ч')}")

async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global next_check_time
    deleted_data = load_deleted()
    if next_check_time:
        remaining = int(next_check_time - datetime.now().timestamp())
        mins, secs = divmod(max(remaining, 0), 60)
        hours, mins = divmod(mins, 60)
        await update.message.reply_text(
            f"📊 Мониторинг: @{USERNAME}\n"
            f"⏱ Интервал: {INTERVAL_HOURS}ч\n"
            f"⏳ До следующей проверки: {hours}ч {mins}м {secs}с\n"
            f"📁 В базе: {len(seen)} репостов\n"
            f"🗑 Удалено всего: {len(deleted_data)}"
        )
    else:
        await update.message.reply_text("Бот запускается...")

seen = load_seen()

async def monitor_loop(app):
    global seen, next_check_time

    if not seen:
        print("Первый запуск, собираем начальные репосты...")
        await app.bot.send_message(CHAT_ID, "🔄 Первый запуск, собираю текущие репосты...")
        initial = await get_reposts(USERNAME)
        seen = {r["id"] for r in initial if r["id"]}
        save_seen(seen)
        await app.bot.send_message(CHAT_ID, f"✅ Бот запущен. Запомнено {len(seen)} репостов @{USERNAME}.")
    else:
        await app.bot.send_message(CHAT_ID, f"✅ Бот перезапущен. В базе {len(seen)} репостов @{USERNAME}.")

    next_check_time = datetime.now().timestamp() + INTERVAL_HOURS * 3600

    while True:
        await asyncio.sleep(10)
        if not is_checking and datetime.now().timestamp() >= next_check_time:
            asyncio.create_task(do_check(app=app))

async def main():
    request = HTTPXRequest(proxy=PROXY) if PROXY else HTTPXRequest()
    app = Application.builder().token(TELEGRAM_TOKEN).request(request).build()
    app.add_handler(CommandHandler("refresh", cmd_refresh))
    app.add_handler(CommandHandler("time", cmd_time))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CallbackQueryHandler(callback_time, pattern="^time_"))

    await app.initialize()
    await app.start()
    await app.updater.start_polling()

    await monitor_loop(app)

if __name__ == "__main__":
    asyncio.run(main())
